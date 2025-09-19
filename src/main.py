# IMPORTANT #
# -------------------------------- #

# Toggle for True to test, and False to ship
test = True

# IMPORTANT #
# -------------------------------- #

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
from datetime import datetime
from sql import get_data
import os
import numpy as np
from helpers import load_flags, save_flags, send_email

df = get_data()
df = df[df['Missed_Flag'] ==  1]
df['YearsOfService'] = pd.to_numeric(df['YearsOfService'], errors='coerce').fillna(0).astype(int)

grouped_data = (
    df.groupby('Employee_ID')
    .agg(
        Last_Name=('Last_Name', 'first'),
        First_Name=('First_Name', 'first'),
        Shift_Assignment=('Shift_Assignment', 'first'),
        Rank=('Rank', 'first'),
        YearsOfService=('YearsOfService', 'first'),
        Shift_Hours=('Shift_Hours', 'sum')
    )
    .reset_index()
)

grouped_data['Shift_Hours'] = (
    pd.to_numeric(grouped_data['Shift_Hours'], errors='coerce')
      .fillna(0)
      .round(2)
)

cutoff = np.where(grouped_data["YearsOfService"] >= 5, 1104, 552)
grouped_data["Remaining Hours"]  = (cutoff - grouped_data["Shift_Hours"]).round(2)
grouped_data = grouped_data.sort_values("Remaining Hours", ascending=True)


if test:
    output_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
    os.makedirs(output_folder, exist_ok=True)
    file_path = os.path.join(output_folder, "Sick_and_Injured.xlsx")
else:
    file_path = r"\\cor.local\Fire\Department\Admin Projects\Payroll\Sick_and_Injured_Report\Sick_and_Injured.xlsx"

wb = Workbook()
ws = wb.active

# Add the text and timestamp in A1 and A2
ws["A1"] = "Report updated on:"
ws["A1"].font = Font(bold=True)
timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
ws["B1"] = timestamp
ws["B1"].font = Font(bold=True)

headers = list(grouped_data.columns)

cache_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "cache")
os.makedirs(cache_dir, exist_ok=True)

yellow_path = os.path.join(cache_dir, "yellows.json")
red_path    = os.path.join(cache_dir, "reds.json")

yellow_flags = load_flags(yellow_path) 
red_flags = load_flags(red_path) 

new_yellows = []
new_reds = []

if "Date Yellow" not in headers:
    headers.append("Date Yellow")
    ws.cell(row=3, column=len(headers), value="Date Yellow")

if "Date Red" not in headers:
    headers.append("Date Red")
    ws.cell(row=3, column=len(headers), value="Date Red")


# Write the result data into the worksheet, starting from row 4
for row_num, row_data in enumerate(grouped_data.itertuples(index=False), 4):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_num, column=col_num, value=value)

        years_of_service = row_data[headers.index("YearsOfService")]
        employee_id = int(row_data[headers.index("Employee_ID")])
        name = f"{row_data[headers.index('Last_Name')]}, {row_data[headers.index('First_Name')]}"
        today = datetime.today().strftime("%Y-%m-%d")

        
        # Apply conditional formatting based on the conditions
        if headers[col_num - 1] == "Shift_Hours":
            cell.number_format = "0.00"
            total_hours = value
            cell_rem = ws.cell(row=row_num, column=headers.index("Remaining Hours") + 1)
            cell_rem.number_format = "0.00"

            # --- YELLOW condition ---
            if (years_of_service >= 5 and 864 <= total_hours < 1104) or \
               (years_of_service < 5 and 312 <= total_hours < 552):
                
                for i in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=i).fill = PatternFill(
                        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                    )
                rec = next((r for r in yellow_flags if r["employee_id"] == employee_id), None)
                if rec:
                    ws.cell(row=row_num, column=headers.index("Date Yellow") + 1,
                            value=rec.get("date_yellow"))
                else:
                    yellow_flags.append({"employee_id": employee_id, "name": name, "date_yellow": today, "date_red": None})
                    new_yellows.append(name)
                    ws.cell(row=row_num, column=headers.index("Date Yellow") + 1, value = today)
                
            # --- RED condition ---
            elif (years_of_service >= 5 and total_hours >= 1104) or \
                 (years_of_service < 5 and total_hours >= 552):
                
                for i in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=i).fill = PatternFill(
                        start_color="FF0000", end_color="FF0000", fill_type="solid"
                    )
                rec = next((r for r in red_flags if r["employee_id"] == employee_id), None)
                if rec:
                    ws.cell(row=row_num, column=headers.index("Date Red") + 1,
                            value=rec.get("date_red"))
                else:
                    red_flags.append({
                        "employee_id": employee_id,
                        "name": name,
                        "date_yellow": None,
                        "date_red": today
                    })
                    new_reds.append(name)
                    ws.cell(row=row_num, column=headers.index("Date Red") + 1, value=today)
      
        # Center align the cell content
        cell.alignment = Alignment(horizontal='left')

for col_num, header in enumerate(headers, 1):
    col_letter = get_column_letter(col_num)
    ws[f"{col_letter}3"] = header
    ws[f"{col_letter}3"].font = Font(bold=True)

save_flags(yellow_path, yellow_flags)
save_flags(red_path, red_flags)

# Auto-fit column widths
for col in ws.columns:
    max_length = 0
    col_letter = get_column_letter(col[0].column)  # Get the column letter
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = (max_length + 2)  # Adding a little extra padding
    ws.column_dimensions[col_letter].width = adjusted_width

# Save the workbook
wb.save(file_path)

if new_yellows:
    yellow_bullets = "<br>".join(f"&nbsp;&nbsp;&nbsp;• &nbsp;{m}" for m in new_yellows)
else:
    yellow_bullets = "&nbsp;&nbsp;&nbsp;• &nbsp;There are no yellow flags."

if new_reds:
    red_bullets = "<br>".join(f"&nbsp;&nbsp;&nbsp;• &nbsp;{m}" for m in new_reds)
else:
    red_bullets = "&nbsp;&nbsp;&nbsp;• &nbsp;There are no red flags."

test_mode_text = ("""
<b>* TEST *</b>
<p>Hi Dan, I think this report is ready ship.  Please take a look at the attached report and also take a look at the documentation in the github repo:</p>
<p>https://github.com/RochesterFireDev/sick_bank_flag</p>
<p>-------------------------------------</p>                  
""" if test else "")

body = f"""      
{test_mode_text}
<p>Hello,</p>

<p>New members have been flagged for their sick-bank hours:</p>

<p><b>YELLOW FLAGS</b>—The following members have approximately 240 hours (~10 working shifts) before going into a no pay status:</p>
<p>{yellow_bullets}</p>

<p><b>RED FLAGS</b>—The following members have exhausted allocated sick hours and could go into a no pay status:</p>
<p>{red_bullets}</p>

The updated sick and injured report is attached.  It can also be viewed here: 
<a href="file:///G:/Admin%20Projects/Payroll/Sick_and_Injured_Report/">
G:\\Admin Projects\\Payroll\\Sick_and_Injured_Report
</a></p>

<p>Rochester Fire Department</p>
<p>(<i>This is an automated message</i>)</p>
"""
send_email(file_path=file_path, body = body, test=test)