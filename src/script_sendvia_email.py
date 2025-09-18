from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import os
import pandas as pd
import pyodbc
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime

# -------------------------
# Database connection
# -------------------------
server = 'RFD-REDNMX-DB'
database = 'RedNMX'
conn = pyodbc.connect(f'DRIVER={{SQL Server}};SERVER={server};DATABASE={database};Trusted_Connection=yes;')

# -------------------------
# SMTP / Email settings
# -------------------------
SMTP_SERVER = "smtp.cor.local"
SMTP_PORT = 25
FROM_EMAIL = "rfdutilsvc@cityofrochester.gov"
TO_EMAILS = ["Sarah.Ruekberg@CityofRochester.Gov","Tracy.Brown@CityofRochester.Gov"] 
#TO_EMAILS = ["willis.sontheimer@cityofrochester.gov","wpsontheimer@gmail.com"] 


# SQL Query
query = '''
select 
    datetimestart datetimeshift,
    shiftlength,
    schdtypecode schdtype_code,
    unitnum,
    schdtypedescr schdtype_descr,
    schdshiftnamedescr shift_descr,
    case 
        when schdlocdescr is null then 'LINE DIVISION' 
        else schdlocdescr 
    end as location,
    schdrankdescr rank,
    lname,
    fname,
    perscode 
from vwschdhist
WHERE (datetimestart >= DATEFROMPARTS(
                        CASE 
                            WHEN MONTH(GETDATE()) >= 7 THEN YEAR(GETDATE()) 
                            ELSE YEAR(GETDATE()) - 1 
                        END, 
                        7, 
                        1)
    AND datetimestart <= GETDATE())
    and 
    (schdtypedescr like '%over%' or schdtypeid = 89)
order by datetimestart
'''

data = pd.read_sql(query, conn)
conn.close()


# Create an Excel workbook
workbook = Workbook()
sheet = workbook.active
sheet.title = "Overtime Report"

header_font = Font(bold=True)

for col_idx, column_name in enumerate(data.columns, start=1):
    cell = sheet.cell(row=1, column=col_idx, value=column_name)
    cell.font = header_font  # Apply bold font

# Write data rows
for row_idx, row in enumerate(data.itertuples(index=False), start=2):
    for col_idx, value in enumerate(row, start=1):
        sheet.cell(row=row_idx, column=col_idx, value=value)

# Autofit column widths
for col_idx, column_cells in enumerate(sheet.columns, start=1):
    max_length = 0
    column_letter = column_cells[0].column_letter  # Get column letter
    for cell in column_cells:
        try:
            if cell.value:  # Check if cell value is not None
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = max_length + 2  # Add padding for better visibility
    sheet.column_dimensions[column_letter].width = adjusted_width

# -------------------------
# Email helper (single recipient, single attachment)
# -------------------------
def send_email(smtp_server, smtp_port, from_email, to_recipients, subject, body_html, attachment_path, cc=None, bcc=None):
    """
    to_recipients, cc, bcc can be a single string or a list of strings.
    """
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email import encoders
    import os, smtplib

    def as_list(x):
        if not x:
            return []
        return list(x) if isinstance(x, (list, tuple, set)) else [x]

    to_list  = [a.strip() for a in as_list(to_recipients)]
    cc_list  = [a.strip() for a in as_list(cc)]
    bcc_list = [a.strip() for a in as_list(bcc)]

    if not to_list:
        print("No recipient found. Skipping email.")
        return

    msg = MIMEMultipart()
    msg["From"] = from_email
    msg["To"] = ", ".join(to_list)
    if cc_list:
        msg["Cc"] = ", ".join(cc_list)
    msg["Subject"] = subject
    msg.attach(MIMEText(body_html, "html"))

    if isinstance(attachment_path, str) and os.path.exists(attachment_path):
        with open(attachment_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{os.path.basename(attachment_path)}"')
        msg.attach(part)
    else:
        print(f"Attachment not found: {attachment_path}")

    # SMTP envelope recipients = To + Cc + Bcc
    all_rcpts = to_list + cc_list + bcc_list

    with smtplib.SMTP(smtp_server, smtp_port) as s:
        s.sendmail(from_email, all_rcpts, msg.as_string())
        print(f"Email sent to {', '.join(all_rcpts)} with attachment: {attachment_path}")

# -------------------------
# Save file and send email
# -------------------------
output_filename = f"Overtime_Report_{datetime.now():%Y%m%d}.xlsx"
workbook.save(output_filename)

email_subject = f"RFD Overtime Data — {datetime.now():%Y-%m-%d}"
email_body = """Hello,<br><br>Attached is all overtime data for the Rochester Fire Department for the current Fiscal Year.<br><br>Thank you,<br><br>- Rochester Fire Department<br>(This is an automated message.  Please do not reply to this email.)"""

send_email(
    smtp_server=SMTP_SERVER,
    smtp_port=SMTP_PORT,
    from_email=FROM_EMAIL,
    to_recipients=TO_EMAILS,  # list works
    subject=email_subject,
    body_html=email_body,
    attachment_path=output_filename,
    # cc=["cc1@cityofrochester.gov"],       # optional
    # bcc=["bcc1@cityofrochester.gov"],     # optional
)